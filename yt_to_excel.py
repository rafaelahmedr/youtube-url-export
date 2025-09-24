import json
import subprocess
import sys
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def fetch_videos(channel_url: str):
    """Fetch videos from a YouTube channel using yt-dlp."""
    result = subprocess.run(
        ["yt-dlp", "-j", "--flat-playlist", channel_url],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        encoding="utf-8",
    )

    videos = []
    for line in result.stdout.splitlines():
        try:
            v = json.loads(line)
            title = v.get("title", "")
            video_id = v.get("id", "")
            url = f"https://www.youtube.com/watch?v={video_id}"
            videos.append((title, url))
        except json.JSONDecodeError:
            continue

    return videos

def save_to_excel(videos, output_file="videos.xlsx"):
    """Save video titles and URLs to an Excel file with clickable links."""
    wb = Workbook()
    ws = wb.active
    ws.title = "YouTube Videos"

    # Headers
    ws.append(["Title", "URL"])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Add rows
    for title, url in videos:
        ws.append([title, url])

    # Make URLs clickable
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        cell = row[0]
        cell.hyperlink = cell.value
        cell.style = "Hyperlink"

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 80)

    wb.save(output_file)
    print(f"Done! Saved as {output_file}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python yt_to_excel.py <channel_url>")
        sys.exit(1)

    channel_url = sys.argv[1]
    videos = fetch_videos(channel_url)
    save_to_excel(videos)
